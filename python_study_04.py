'''
7. 엑셀 파일 읽고 쓰는 방법
- OpenPyXL 라이브러리 설치
pip install openpyxl
- 엑셀 파일 쓰기
'''
import openpyxl

wb = openpyxl.Workbook() #워크북 인스턴스 객체 생성
sheet = wb.active #현재 활성화된 워크시트 객체 가져옴
sheet.title = '회원정보' #기본값 Sheet -> 회원정보

#헤더 정보
header_titles = ['이름', '전화번호']
for idx, title in enumerate(header_titles):
    sheet.cell(row=1, column=idx+1, value=title)

#내용
members = [('홍길동', '010-1111-1111'), ('이영희', '010-2222-2222')]
row_num = 2 #1번째 row는 타이틀 위치
for r, member in enumerate(members): #회원정보 목록 탐색
    for c, v in enumerate(member): #이름, 전화번호 컬럼 탐색
        sheet.cell(row=row_num+r, column=c+1, value=v)
wb.save('./members.xlsx')
wb.close()

'''
- 엑셀 파일 읽기
'''
wb = openpyxl.load_workbook('./members.xlsx')
sheet = wb['회원정보']

#워크시트 내 모든 row 데이터 탐색, 초기 탐색 위치는 2
for row in sheet.iter_rows(min_row=2):
    for cell in row:
        print(cell.value)
wb.close()

